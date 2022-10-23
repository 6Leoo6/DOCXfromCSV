from io import BytesIO
from random import choice
from string import ascii_letters, digits
from xml.dom import minidom
from zipfile import ZipFile

import boto3
from app.numtoletter import int_to_letter
from openpyxl import load_workbook


async def convertXLSX(excel_file, model_file):
    # Read the bytes of the files uploaded
    excel_bytes = BytesIO(await excel_file.read())
    model_bytes = BytesIO(await model_file.read())

    # Parse the Excel
    excel = load_workbook(excel_bytes).active

    # Separate the first row (Headers) from the others
    excel_headers = [excel.cell(row=1, column=col).value for col in range(1, excel.max_column+1)]
    excel_rows = [[str(excel.cell(row=row, column=col).value) for col in range(1, excel.max_column+1)] for row in range(2, excel.max_row+1)]

    #Create a new empty zip in memory
    zip_bytes = BytesIO()
    #Open it for writing
    with ZipFile(zip_bytes, 'w') as zip:
        #Loop through the rows of the excel
        for row in excel_rows:
            files = []
            #Create an empty XML file in memory
            doc_xml = BytesIO()
            #Define the default filename
            filename = row[0]
            
            #Open the document model as a zip
            with ZipFile(model_bytes, 'r') as doc:
                #Loop through the files of the document
                for file in doc.filelist:
                    fname = file.filename
                    #If the file is the main file then save it for later use and continue
                    if fname == 'word/document.xml':
                        doc_xml = doc.read(fname).decode('utf-8')
                        continue
                    #Store all the other files temporarly to prevent creating duplicates
                    files.append([fname, doc.read(fname)])

            #Loop through the columns
            #Reverse the row to prevent the replacement of VARAB by VARA
            #(Loop in descending order by the length of the id's)
            for col in reversed(range(len(excel_headers))):
                # Replace the id with the corresponding value
                current_id = f'XVAR{int_to_letter(col+1)}'
                doc_xml = doc_xml.replace(current_id, row[col])

            #Parse the replaced XML
            xml = minidom.parseString(doc_xml)
            #Get the body of the file
            body = xml.firstChild.firstChild
            first_p = ''
            try:
                #Reconstruct the first paragraph by looping through it and adding the values
                for child in body.firstChild.childNodes:
                    if child.tagName != 'w:r':
                        continue
                    for c in child.childNodes:
                        if c.tagName != 'w:t':
                            continue
                        first_p += c.firstChild.nodeValue
                #If the text start with and ends with a square bracket
                if first_p.startswith('[') and first_p.endswith(']'):
                    #Set the parsed paragraph as the filename
                    filename = first_p.strip('[]')
                    #Remove the paragraph from the document
                    body.removeChild(body.firstChild)
            except (AttributeError, TypeError): #If the first element is an image, there will be an error
                pass
            #Convert back to text
            doc_xml = xml.toxml()

            #Create an empty doc
            doc_bytes = BytesIO()
            #Read the doc as a zip
            with ZipFile(doc_bytes, 'w') as doc:
                #Put back the files to the zip
                for fname, file in files:
                    if fname != 'word/document.xml':
                        doc.writestr(fname, file)
                #Put back the modified XML
                doc.writestr('word/document.xml', doc_xml)

            #Save the file as a docx to the main zip
            zip.writestr(f'{filename}.docx', doc_bytes.getvalue())

    zip_bytes.seek(0)

    # Upload to s3
    return upload_to_s3(zip_bytes)


def upload_to_s3(file: BytesIO):
    bucket_name = 'uploaded-files-docx-from-excel'

    s3 = boto3.client("s3")

    existing_names = [key['Key'] for key in s3.list_objects(Bucket=bucket_name)['Contents']]
    while True:
        obj_name = ''.join([choice(ascii_letters + digits) for _ in range(15)]) + '.zip'
        if obj_name not in existing_names:
            break

    s3.upload_fileobj(file, bucket_name, obj_name)

    url = s3.generate_presigned_url('get_object',
                                    Params={'Bucket': bucket_name,
                                            'Key': obj_name},
                                    ExpiresIn=3600
    )

    return url